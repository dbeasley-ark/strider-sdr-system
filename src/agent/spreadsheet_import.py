"""Parse prospect lists from CSV or XLSX for batch sales runs.

Detects a company column by header name; optional domain column.
No dependency on LLM runtime or agent config.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path

_COMPANY_HEADER_PRIORITY: tuple[str, ...] = (
    "company",
    "company name",
    "organization",
    "organisation",
    "org",
    "prospect",
    "account",
    "legal name",
    "name",
    "domain",
    "website",
    "url",
)

_DOMAIN_HEADER_PRIORITY: tuple[str, ...] = (
    "domain",
    "company domain",
    "website",
    "web site",
    "url",
    "site",
)


def _norm_header(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def _pick_column(
    headers: list[str],
    priority: tuple[str, ...],
    *,
    exclude_idx: int | None = None,
) -> int | None:
    normalized = [_norm_header(h) for h in headers]
    for want in priority:
        for i, h in enumerate(normalized):
            if exclude_idx is not None and i == exclude_idx:
                continue
            if h == want:
                return i
    for want in priority:
        for i, h in enumerate(normalized):
            if exclude_idx is not None and i == exclude_idx:
                continue
            if h.startswith(want) or want in h:
                return i
    return None


@dataclass(frozen=True)
class ParsedSheet:
    headers: list[str]
    company_column: int
    domain_column: int | None
    rows: list[tuple[str, str | None]]  # (company, domain_hint)


class SpreadsheetParseError(ValueError):
    pass


def _read_rows_csv(raw: bytes) -> tuple[list[str], list[list[str]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_in = list(reader)
    if not rows_in:
        raise SpreadsheetParseError("The file is empty.")
    headers = [c.strip() for c in rows_in[0]]
    data = [[(c or "").strip() for c in row] for row in rows_in[1:]]
    return headers, data


def _read_rows_xlsx(raw: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SpreadsheetParseError(
            "XLSX support requires openpyxl. Install with: pip install openpyxl"
        ) from e

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            raise SpreadsheetParseError("The spreadsheet has no header row.")
        headers = [("" if c is None else str(c)).strip() for c in header_row]
        data: list[list[str]] = []
        for row in rows_iter:
            data.append([("" if c is None else str(c)).strip() for c in row])
    finally:
        wb.close()
    return headers, data


def parse_prospect_spreadsheet(
    raw: bytes,
    *,
    filename: str,
    company_column: str | None = None,
    domain_column: str | None = None,
    max_rows: int = 500,
) -> ParsedSheet:
    """Parse CSV or XLSX. Headers must be in the first row."""
    suffix = Path(filename).suffix.lower()
    if suffix in {".csv", ".txt"}:
        headers, data = _read_rows_csv(raw)
    elif suffix in {".xlsx", ".xlsm"}:
        headers, data = _read_rows_xlsx(raw)
    else:
        raise SpreadsheetParseError(
            f"Unsupported file type {suffix!r}. Use .csv or .xlsx."
        )

    if not any(h for h in headers):
        raise SpreadsheetParseError("The header row is empty.")

    company_idx: int | None = None
    if company_column:
        want = _norm_header(company_column)
        for i, h in enumerate(headers):
            if _norm_header(h) == want:
                company_idx = i
                break
        if company_idx is None:
            raise SpreadsheetParseError(
                f"No column matching company column {company_column!r}."
            )
    else:
        company_idx = _pick_column(headers, _COMPANY_HEADER_PRIORITY)
        if company_idx is None:
            raise SpreadsheetParseError(
                "Could not detect a company column. "
                "Name a column Company, Organization, Prospect, or Domain, "
                "or pass company_column explicitly."
            )

    domain_idx: int | None = None
    if domain_column:
        want = _norm_header(domain_column)
        for i, h in enumerate(headers):
            if _norm_header(h) == want:
                domain_idx = i
                break
        if domain_idx is None:
            raise SpreadsheetParseError(
                f"No column matching domain column {domain_column!r}."
            )
    else:
        domain_idx = _pick_column(
            headers, _DOMAIN_HEADER_PRIORITY, exclude_idx=company_idx
        )

    out: list[tuple[str, str | None]] = []
    for row in data:
        if company_idx >= len(row):
            continue
        company = row[company_idx].strip()
        if not company:
            continue
        domain: str | None = None
        if domain_idx is not None and domain_idx < len(row):
            d = row[domain_idx].strip()
            domain = d or None
        out.append((company, domain))
        if len(out) >= max_rows:
            break

    if not out:
        raise SpreadsheetParseError("No data rows with a non-empty company value.")

    return ParsedSheet(
        headers=headers,
        company_column=company_idx,
        domain_column=domain_idx,
        rows=out,
    )
